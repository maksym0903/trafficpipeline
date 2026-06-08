"""Report, enrichment, and indexability helpers for NextRce."""
import csv
import json
import os
import re
import shlex
from urllib.parse import urljoin

from nextrce_config import (
    ENRICH_COMMANDS,
    ENRICH_TIMEOUT,
    FIND_HTTP_TIMEOUT,
    INDEXABLE_FILE,
    REPORT_COLUMNS,
    STACK_DOMAIN_FILES,
    STACK_FEATURE_FILES,
    STACK_OUTPUT_FILES,
    Colors,
)
from nextrce_http import extract_url, http_get

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

def project_root():
    """Repo root: parent of tools/ when run as tools/nextrce.py, else cwd."""
    here = os.path.dirname(os.path.abspath(__file__))
    if os.path.basename(here) == 'tools':
        return os.path.dirname(here)
    return os.getcwd()


def find_default_target_list():
    """Default pipeline input: urls/target.txt under project root or cwd."""
    for base in (os.getcwd(), project_root()):
        path = os.path.join(base, 'urls', 'target.txt')
        if os.path.isfile(path):
            return path
    return None


def load_target_file(path):
    """Load URLs from a list file; skip blanks and # comments."""
    targets = []
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            url = extract_url(line) or line
            targets.append(url)
    return targets


def empty_report_fields():
    return {col: '' for col in REPORT_COLUMNS[3:]}


def flatten_for_csv(text):
    """Collapse multiline command output into one CSV-safe line."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return ' | '.join(lines)


def is_shell_error(line):
    lower = line.lower().strip()
    if not lower:
        return True
    if lower.startswith('/bin/sh:') or lower.startswith('sh:'):
        return True
    if 'cannot fork' in lower or "can't fork" in lower:
        return True
    if 'not found' in lower and (': 1:' in lower or ': 127:' in lower):
        return True
    return False


def clean_shell_output(text):
    """Drop shell stderr noise; keep real command output only."""
    lines = [line.strip() for line in text.splitlines()
             if line.strip() and not is_shell_error(line)]
    return ' | '.join(lines)


def filter_app_package_paths(text):
    """Keep real app package.json paths; drop node_modules/cache noise."""
    return ' | '.join(parse_package_json_paths(text))


def parse_package_json_paths(text):
    skip_parts = ('/node_modules/', '/.cache/', '/yarn/v6/', '/.npm/')
    seen = set()
    results = []
    for line in text.splitlines():
        path = line.strip()
        if not path.endswith('package.json'):
            continue
        if any(part in path for part in skip_parts):
            continue
        if path in seen:
            continue
        seen.add(path)
        results.append(path)
        if len(results) >= 20:
            break
    return results


def is_priority_package_path(path):
    """Paths that trigger package.json inspection."""
    if path == '/app/package.json':
        return True
    if re.match(r'^/var/www/[^/]+/package\.json$', path):
        return True
    return False


def select_inspect_paths(paths, webroot=''):
    """Pick package.json paths worth inspecting for app stack type."""
    inspect = []
    for path in paths:
        if is_priority_package_path(path):
            inspect.append(path)
    if webroot:
        candidate = webroot.rstrip('/') + '/package.json'
        if candidate in paths and candidate not in inspect:
            inspect.append(candidate)
    return inspect


def classify_package_json(content):
    """Classify app stack from package.json contents."""
    if not content or not content.strip():
        return 'Static'
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        return 'Static'

    deps = {}
    for key in ('dependencies', 'devDependencies', 'peerDependencies'):
        deps.update(data.get(key) or {})

    content_lower = content.lower()
    dep_keys = ' '.join(deps.keys()).lower()

    if 'next' in deps:
        return 'Next.js'
    if any('wordpress' in k.lower() for k in deps) or '@wordpress/' in content_lower:
        return 'WordPress'
    if 'laravel-vite-plugin' in deps or '@laravel/' in dep_keys or 'laravel' in content_lower:
        return 'Laravel'
    if 'react' in deps or 'react-dom' in deps:
        return 'React'
    return 'Static'


def parse_os_release(text):
    for line in text.splitlines():
        if line.startswith('PRETTY_NAME='):
            return line.split('=', 1)[1].strip().strip('"')
    for line in text.splitlines():
        if line.startswith('NAME='):
            return line.split('=', 1)[1].strip().strip('"')
    return flatten_for_csv(text)


def format_enrich_value(column, output):
    if not output:
        return ''
    if column == 'os':
        return parse_os_release(output) or clean_shell_output(output)
    if column == 'server_country':
        code = clean_shell_output(output).strip().upper()
        return code if len(code) == 2 and code.isalpha() else ''
    return clean_shell_output(output)


def build_report_rows(targets, report_rows):
    """Build report data rows (header + one row per target)."""
    rows = [REPORT_COLUMNS]
    for url in targets:
        row = report_rows.get(url)
        if not row:
            row = {'status': 'failed', 'access': 'no', **empty_report_fields()}
        rows.append(
            [url, row['status'], row['access']]
            + [row.get(col, '') for col in REPORT_COLUMNS[3:]]
        )
    return rows


def write_stack_files_from_rows(rows, results_dir):
    """Write stack/feature txt files from report rows (same data as report.csv)."""
    if len(rows) < 2:
        return []

    col = {name: idx for idx, name in enumerate(rows[0])}
    buckets = {fname: [] for fname in STACK_OUTPUT_FILES}

    for row in rows[1:]:
        if row[col['status']] != 'success':
            continue
        url = row[col['url']]
        domain_type = row[col['domain_type']]
        if domain_type in STACK_DOMAIN_FILES:
            buckets[STACK_DOMAIN_FILES[domain_type]].append(url)
        if row[col['nodejs']].strip():
            buckets[STACK_FEATURE_FILES['nodejs']].append(url)
        if row[col['docker']].strip():
            buckets[STACK_FEATURE_FILES['docker']].append(url)
        if row[col['indexable']].strip().lower() == 'yes':
            buckets[INDEXABLE_FILE].append(url)

    os.makedirs(results_dir, exist_ok=True)
    written = []
    for fname in STACK_OUTPUT_FILES:
        path = os.path.join(results_dir, fname)
        with open(path, 'w', encoding='utf-8') as f:
            for url in buckets[fname]:
                f.write(url + '\n')
        if buckets[fname]:
            written.append(f"{fname} ({len(buckets[fname])})")
    return written


def export_stack_files_from_csv(csv_path, results_dir=None):
    """Regenerate stack txt files from an existing report.csv."""
    results_dir = results_dir or os.path.dirname(os.path.abspath(csv_path))
    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    return write_stack_files_from_rows(rows, results_dir)


def load_report_from_csv(csv_path):
    """Load report.csv back into targets list + report_rows dict."""
    targets = []
    report_rows = {}
    with open(csv_path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            url = row['url']
            targets.append(url)
            report_rows[url] = {col: row.get(col, '') for col in REPORT_COLUMNS if col != 'url'}
    return targets, report_rows


def check_indexable(opener, url, timeout=10):
    """
    SEO indexability check (public HTTP, no RCE):
    - homepage HTTP 200
    - no X-Robots-Tag: noindex
    - no meta robots noindex
    - robots.txt does not disallow all for User-agent: *
    """
    try:
        headers, html, status = http_get(opener, url, timeout)
    except Exception:
        return 'no'

    if status != 200:
        return 'no'

    robots_tag = (headers.get('X-Robots-Tag') or headers.get('x-robots-tag') or '').lower()
    if 'noindex' in robots_tag:
        return 'no'

    snippet = html[:100000].lower()
    if re.search(r'<meta[^>]+name=["\']robots["\'][^>]+noindex', snippet):
        return 'no'
    if re.search(r'<meta[^>]+content=["\'][^"\']*noindex', snippet):
        return 'no'

    try:
        _, robots_body, robots_status = http_get(opener, urljoin(url, '/robots.txt'), timeout)
        if robots_status == 200:
            body = robots_body.lower()
            for block in re.split(r'(?=user-agent:)', body):
                if re.match(r'user-agent:\s*\*', block.strip()):
                    if re.search(r'disallow:\s*/\s*$', block, re.MULTILINE):
                        return 'no'
                    break
    except Exception:
        pass

    return 'yes'


def write_report_xlsx(path, rows):
    """Write formatted Excel table (.xlsx). Requires openpyxl."""
    if not _HAS_OPENPYXL or len(rows) < 2:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    last_col = get_column_letter(len(REPORT_COLUMNS))
    last_row = len(rows)
    table = Table(
        displayName='ScanReport',
        ref=f'A1:{last_col}{last_row}',
    )
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = 'A2'

    for col_idx in range(1, len(REPORT_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(rows[r][col_idx - 1] or '')) for r in range(len(rows)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    wb.save(path)
    return True


def print_banner():
    print(rf"""
{Colors.CYAN}    _   __          __  ____            
   / | / /__  _  __/ /_/ __ \________   
  /  |/ / _ \| |/_/ __/ /_/ / ___/ _ \  
 / /|  /  __/>  </ /_/ _, _/ /__/  __/  
/_/ |_/\___/_/|_|\__/_/ |_|\___/\___/   
{Colors.RESET}
{Colors.BOLD}   Next.js RSC Exploit Tool (CVE-2025-55182){Colors.RESET}
{Colors.GREY}   Mass Scanner & Pipeline Edition (v3.3 · no-deps){Colors.RESET}

{Colors.RED}   >> OPERATOR: MITSEC ( x.com/ynsmroztas ){Colors.RESET}
""")
